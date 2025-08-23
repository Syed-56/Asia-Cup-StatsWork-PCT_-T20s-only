import kagglehub
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

# Step 1: Download dataset
path = kagglehub.dataset_download("hasibalmuzdadid/asia-cup-cricket-1984-to-2022")

# Step 2: Load CSVs into DataFrames (these will be importable too)
matches = pd.read_csv(f"{path}/asiacup.csv")
matches = matches[matches["Format"] == "T20I"]

champions = pd.read_csv(f"{path}/champion.csv")
champions = champions[champions["Year"].isin([2016, 2022])]

batsmanDataT20 = pd.read_csv(f"{path}/batsman data t20i.csv")
bowlerDataT20 = pd.read_csv(f"{path}/bowler data t20i.csv")
wkpDataT20 = pd.read_csv(f"{path}/wicketkeeper data t20i.csv")

# Step 3: Ensure Data folder exists
os.makedirs("Data", exist_ok=True)
excel_path = "Data/asia_cup_t20_analysis.xlsx"


def format_excel(path: str):
    """Apply header style, column auto-width, freeze header."""
    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Bold headers + fill color
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Auto column width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(path)


if __name__ == "__main__":
    # Step 4: Write to Excel
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        matches.to_excel(writer, sheet_name="Matches", index=False)
        champions.to_excel(writer, sheet_name="Champions", index=False)
        batsmanDataT20.to_excel(writer, sheet_name="Batsman", index=False)
        bowlerDataT20.to_excel(writer, sheet_name="Bowler", index=False)
        wkpDataT20.to_excel(writer, sheet_name="Wicketkeeper", index=False)

    # Step 5: Apply formatting
    format_excel(excel_path)

    print("✅ Excel file created with formatting!")
    print(matches.head())